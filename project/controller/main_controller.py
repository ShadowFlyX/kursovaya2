import sys
import os

import PySide6.QtCore
import PySide6.QtWidgets
import openpyxl.cell

project_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../"))   # Получаем путь к каталогу проекта

sys.path.append(project_dir)                                            # Добавляем каталог проекта в PYTHONPATH
import PySide6
from PySide6.QtWidgets import QApplication, QMainWindow
from view.main_view import Ui_MainWindow
import sqlalchemy as db
from sqlalchemy.orm import sessionmaker
from models.schedule_model import ScheduleModel
import openpyxl
from openpyxl.styles import (Border, Side, Alignment, Font)
from controller.process_connection import *


def run_application():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


conntection_string = create_connection_string(read_settings("project/settings.txt"))

engine = db.create_engine(conntection_string)                          

dirname = os.path.dirname(PySide6.__file__)                            

plugin_path = os.path.join(dirname, "plugins", "platforms")
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path


Session = sessionmaker(bind=engine)                                   

class Controller:
    COUNT_STUDY_DAYS = 6

    def __init__(self, view: Ui_MainWindow):
        self.view = view
        self.session = Session()
        self.model = ScheduleModel()
        self.view.pushButton.clicked.connect(self.generate_file)
        self.view.comboBox.addItems(self.get_faculty_data())
        self.default_cell_aligment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.default_cell_font = Font("Times New Roman", 10)
        self.default_day_aligment = Alignment(vertical="center", horizontal="center", text_rotation=90)
        self.default_day_font = Font("Times New Roman", 14, bold=True)
        self.default_header_font = Font("Times New Roman", 20, bold=True)
        self.default_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    def get_faculty_data(self):
        return self.model.get_faculties(self.session)

    @staticmethod
    def showError(title="Error", message="Message!"):
        error = PySide6.QtWidgets.QErrorMessage()
        error.showMessage(message)
        error.setWindowTitle(title)
        error.exec()

    @staticmethod
    def get_lesson_string(data):
        return f"{data.discipline}\n{data.teacher}\n{data.classroom}"

    def generate_file(self):
        
        faculty = self.view.comboBox.currentText() 
        
        if faculty is None:
            self.showError("Error", "Select faculty!")
            return
        
        semester = int(self.view.comboBox1.currentText())
        self.study_time = self.model.get_study_time(self.session, faculty, semester)

        faculty_dict = self.model.get_faculty_groups(self.session, faculty, semester)
        for course, groups in faculty_dict.items():
            groups[:] = self.sort_groups(groups, self.model.get_all_groups_schedule_by_course(self.session, faculty, semester, course))

        faculty_groups = [group for course_groups in faculty_dict.values() for group in course_groups] 
        
        wb = openpyxl.Workbook()
        ws = wb.active
           
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(faculty_groups) * 2 + 4)
        self.format_data_and_set_value(ws, 1, 1, faculty, self.default_cell_aligment, self.default_header_font, self.default_border)
        
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=2)
        self.format_data_and_set_value(ws, 2, 2, "", border=self.default_border)
        start_column = 3 
        for course, groups in faculty_dict.items():
            ws.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=start_column+len(groups)*2-1)
            self.format_data_and_set_value(ws, 2, start_column, f"{course} курс", self.default_cell_aligment, self.default_header_font, self.default_border)
            start_column += len(groups)*2
        #------------------------------------------------------------------------------------------------------------------
          
        if not self.study_time:
            self.showError("Error", "No data for this faculty and semester")
            return
        
        week_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        count_lessons = len(self.study_time)

        self.write_study_time(ws, 1, 2, week_days, count_lessons)
        
        #------------------------------------------------------------------------------------------------------------------
        
        for row in range(count_lessons*2*self.COUNT_STUDY_DAYS+4):
            ws.row_dimensions[row].height = 40


        column = 3
        for group in faculty_groups:
            ws.merge_cells(start_row=3, start_column=column, end_row=3, end_column=column+1) 
        
            self.format_data_and_set_value(ws, 3, column, group, self.default_cell_aligment, self.default_header_font, self.default_border)

            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(column)].width = 15
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(column+1)].width = 15
            
            week_schedule = self.model.get_schedule_for_week(self.session, group, semester)
            
            week_schedule_dict = self.process_schedule(week_schedule, self.study_time)

            row = 4
            for day in week_schedule_dict:
                for time, data in week_schedule_dict[day].items():  
                    if data:    
                        length = len(data)
                        if data[0].overunderline is not None:
                            if data[0].overunderline.lower() == "над чертой":
                                if data[0].subgroup:
                                    if data[0].subgroup == "1":
                                        self.format_cell_and_set_value(ws, row, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        if length >= 2:
                                            if data[1].overunderline is None:
                                                ws.merge_cells(start_row=row, start_column=column+1, end_row=row+1, end_column=column+1)
                                                self.format_cell_and_set_value(ws, row, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                if length == 3:
                                                    self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].overunderline.lower() == "над чертой":
                                                self.format_cell_and_set_value(ws, row, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                if length == 3:
                                                    if data[2].subgroup is None:
                                                        ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                                        self.format_cell_and_set_value(ws, row+1, column, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    elif data[2].subgroup == "1":
                                                        self.format_cell_and_set_value(ws, row+1, column, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    elif data[2].subgroup == "2":
                                                        self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                elif length == 4:
                                                    self.format_cell_and_set_value(ws, row+1, column, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    self.format_cell_and_set_value(ws, row+1, column+1, data[3], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].overunderline.lower() == "под чертой":
                                                if length == 2:
                                                    if data[1].subgroup is None:
                                                        ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                                        self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    elif data[1].subgroup == "1":
                                                        self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    elif data[1].subgroup == "2":
                                                        self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                elif length == 3:
                                                    self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                                    self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if data[0].subgroup == "2":
                                        self.format_cell_and_set_value(ws, row, column+1, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        if length == 2:
                                            if data[1].subgroup is None:
                                                ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                                self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].subgroup == "1":
                                                self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].subgroup == "2":
                                                self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        elif length == 3:
                                            self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if data[0].subgroup == "3":
                                        self.format_cell_and_set_value(ws, row, column+1, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        if length == 2:
                                            if data[1].subgroup is None:
                                                ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                                self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].subgroup == "1":
                                                self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].subgroup == "2":
                                                self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        elif length == 3:
                                            self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                else:
                                    ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+1)
                                    self.format_cell_and_set_value(ws, row, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if length == 2:
                                        if data[1].subgroup is None:
                                            ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                            self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        elif data[1].subgroup == "1":
                                            self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        elif data[1].subgroup == "2":
                                            self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    elif length == 3:
                                        self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    else:
                                        ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                        self.format_cell_and_set_value(ws, row+1, column, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                            elif data[0].overunderline.lower() == "под чертой":
                                if length == 1:
                                    if data[0].subgroup is None:
                                        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+1)
                                        self.format_cell_and_set_value(ws, row, column, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
                                        self.format_cell_and_set_value(ws, row+1, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    elif data[0].subgroup == "1":
                                        self.format_cell_and_set_value(ws, row+1, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    elif data[0].subgroup == "2":
                                        self.format_cell_and_set_value(ws, row+1, column+1, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    elif data[0].subgroup == "3":
                                        self.format_cell_and_set_value(ws, row+1, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                elif length == 2:
                                    ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+1)
                                    self.format_cell_and_set_value(ws, row, column, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    self.format_cell_and_set_value(ws, row+1, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                        else:
                            if data[0].subgroup is not None:
                                if data[0].subgroup == "1":
                                    ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column)
                                    self.format_cell_and_set_value(ws, row, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if length == 3:
                                        self.format_cell_and_set_value(ws, row, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        self.format_cell_and_set_value(ws, row+1, column+1, data[2], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    elif length == 2:
                                        if data[1].overunderline is not None:
                                            if data[1].overunderline.lower() == "над чертой":
                                                self.format_cell_and_set_value(ws, row, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                            elif data[1].overunderline.lower() == "под чертой":
                                                self.format_cell_and_set_value(ws, row+1, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                        else:
                                            ws.merge_cells(start_row=row, start_column=column+1, end_row=row+1, end_column=column+1)
                                            self.format_cell_and_set_value(ws, row, column+1, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    else:
                                        ws.merge_cells(start_row=row, start_column=column+1, end_row=row+1, end_column=column+1)
                                        self.format_cell_and_set_value(ws, row, column+1, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                                elif data[0].subgroup == "2":
                                    ws.merge_cells(start_row=row, start_column=column+1, end_row=row+1, end_column=column+1)
                                    self.format_cell_and_set_value(ws, row, column+1, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if length == 1:
                                        ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column)
                                        self.format_cell_and_set_value(ws, row, column, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                                    if length == 2:
                                        self.format_cell_and_set_value(ws, row+1, column, data[1], self.default_cell_aligment, self.default_cell_font, self.default_border)
                                elif data[0].subgroup == "3" and length == 1:
                                    ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column+1)
                                    self.format_cell_and_set_value(ws, row, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                            else:
                                ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column+1)
                                self.format_cell_and_set_value(ws, row, column, data[0], self.default_cell_aligment, self.default_cell_font, self.default_border)
                    else:
                        ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column+1)
                        self.format_cell_and_set_value(ws, row, column, "", self.default_cell_aligment, self.default_cell_font, self.default_border)
                    row += 2
            column += 2  

        #------------------------------------------------------------------------------------------------------------------
        self.merge_similar_lessons(ws, 4, 3, row, column)
        self.write_study_time(ws, column, column+1, week_days, count_lessons)
        ws.merge_cells(start_row=2, start_column=column, end_row=3, end_column=column+1)
        self.format_cell_and_set_value(ws, 2, column, "", border=self.default_border)
        wb.save("additional/test.xlsx")


    def format_cell_and_set_value(self, ws, row, column, data, alignment=None, font=None, border=None):
        cell = ws.cell(row=row, column=column)
        if data:
            cell.value = self.get_lesson_string(data)
        if alignment:
            cell.alignment = alignment
        if font:
            cell.font = font
        if border:
            cell.border = border


    def format_data_and_set_value(self, ws, row, column, data, alignment=None, font=None, border=None):
        cell = ws.cell(row=row, column=column)
        if data:
            cell.value = data
        if alignment:
            cell.alignment = alignment
        if font:
            cell.font = font
        if border:
            cell.border = border


    def sort_groups(self, groups, data):
        dct = {group: set() for group in groups}
        
        schedule = self.process_schedule(data, self.study_time)

        for day in schedule:
            for time in schedule[day]:
                for i, data in enumerate(schedule[day][time]):
                    for j in range(i+1, len(schedule[day][time])):
                        temp = schedule[day][time][j]
                        if self.compare_lessons(data, temp):
                            dct[data.sgroup].add(temp.sgroup)
                            dct[temp.sgroup].add(data.sgroup)
        lst = []
        for k, v in dct.items():
            if k not in lst:
                lst.append(k)
                for group in v:
                    if group not in lst:
                        lst.extend(sorted(v)) 
        return lst
    

    @staticmethod
    def compare_lessons(lesson1, lesson2):
        return lesson1.discipline == lesson2.discipline and lesson1.classroom == lesson2.classroom \
        and lesson1.teacher == lesson2.teacher and lesson1.overunderline == lesson2.overunderline and lesson1.sgroup != lesson2.sgroup


    def write_study_time(self, ws, column_day, column_time, week_days, count_lessons):
        for day in range(self.COUNT_STUDY_DAYS):                    
            row_start = day * count_lessons * 2 + 4
            row_end = row_start + count_lessons * 2 - 1 
            
            ws.merge_cells(start_row=row_start, start_column=column_day, end_row=row_end, end_column=column_day)
            
            self.format_data_and_set_value(ws, row_start, column_day, week_days[day], alignment=self.default_day_aligment, font=self.default_day_font, border=self.default_border)

            for row, time in enumerate(self.study_time):
                time_start_row = row_start + row*2
                ws.merge_cells(start_row=time_start_row, start_column=column_time, end_row=time_start_row+1, end_column=column_time)
                self.format_data_and_set_value(ws, time_start_row, column_time, time.strftime("%#H:%M"), Alignment(vertical="center", text_rotation=90), Font("Times New Roman", 10, italic=True), self.default_border)


    def process_schedule(self, week_schedule, study_time):
        week_schedule_dict = {}

        for day in range(1, self.COUNT_STUDY_DAYS+1):
            week_schedule_dict.setdefault(day, {})
            for time in study_time:
                week_schedule_dict[day].setdefault(time, [])

        for data in week_schedule:
            week_schedule_dict[data.dayofweek][data.timebeg].append(data)
    
        return week_schedule_dict


    def merge_similar_lessons(self, ws, start_row, start_column, end_row, end_column):
        row = start_row
        merged_cells = ws.merged_cells.ranges

        while row <= end_row:
            column = start_column
            while column <= end_column:
                cell1 = ws.cell(row=row, column=column).value
                cell2 = ws.cell(row=row, column=column + 2).value if column + 2 <= end_column else None

                if cell1 and cell1 == cell2:
                    temp_column = column
                    column += 2
                    
                    while column + 2 <= end_column and ws.cell(row=row, column=column + 2).value == cell1:
                        column += 2
                    
                    found = False
                    
                    for merged_range in merged_cells:
                        if (merged_range.min_row == row and merged_range.min_col == column):
                            last_row = merged_range.max_row
                            last_column = merged_range.max_col
                            found = True
                            break
                    
                    if found:
                        for col in range(temp_column, last_column+1, 2):
                            ws.unmerge_cells(start_row=row, start_column=col, end_row=last_row, end_column=col+1)
                        ws.merge_cells(start_row=row, start_column=temp_column, end_row=last_row, end_column=last_column)

                column += 2
            row += 1
     


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.controller = Controller(self.ui)

if __name__ == "__main__":
    run_application()